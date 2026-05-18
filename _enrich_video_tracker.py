"""
Enrich BIO-304-Video-Tracker.xlsx Video Tracker sheet with per-topic URLs.

Adds three columns to the right of the existing 15:
  16 (P): Lecture Page URL  - link to the per-topic one-pager (if mapped)
  17 (Q): Spaced Recall (deep-link)  - opens the cards for this topic
  18 (R): Lab Workbook URL  - the printable workbook

Pulls values from course-content.js. Lab workbook filenames match
build_workbooks.py's slugify rule.

Idempotent: if the columns are already there (header detected), re-fills only.
"""

import json
import os
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "BIO-304-Video-Tracker.xlsx")
BASE_URL = "https://drsrennie-stack.github.io/nonmajors/"

NAVY_FILL = PatternFill(start_color="FF1E3D4C", end_color="FF1E3D4C", fill_type="solid")
WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def slugify(text):
    return re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()


def load_topics():
    proc = subprocess.run(
        ["node", "-e",
         "const w={};new Function('window',require('fs').readFileSync('course-content.js','utf8'))(w);"
         "console.log(JSON.stringify(w.BIO304_COURSE_CONTENT.modules.flatMap("
         "m=>m.topics.map(t=>({id:t.id,title:t.title,day:t.dayInCourse,"
         "lecturePageUrl:t.lecturePageUrl||null,readingUrl:t.readingUrl||null})))))"],
        cwd=HERE, capture_output=True, text=True, check=True,
    )
    return json.loads(proc.stdout)


def main():
    topics = load_topics()
    by_id = {t["id"]: t for t in topics}

    wb = load_workbook(XLSX)
    ws = wb["Video Tracker"]

    # Add or refresh headers in columns 16, 17, 18.
    new_headers = ["Lecture Page", "Spaced Recall (deep-link)", "Lab Workbook"]
    for i, h in enumerate(new_headers, start=16):
        cell = ws.cell(row=5, column=i, value=h)
        cell.fill = NAVY_FILL
        cell.font = WHITE_BOLD
        cell.alignment = HEADER_ALIGN

    # Walk the data rows (row 6 onward) and fill the new columns.
    filled = 0
    skipped = 0
    for r in range(6, ws.max_row + 1):
        tid = ws.cell(row=r, column=3).value  # column C: Topic ID
        if not tid or not isinstance(tid, str) or not tid.startswith("t-"):
            continue
        topic = by_id.get(tid)
        if not topic:
            skipped += 1
            continue

        # Lecture page URL (only set if course-content.js has lecturePageUrl)
        if topic.get("lecturePageUrl"):
            url = BASE_URL + topic["lecturePageUrl"]
            cell = ws.cell(row=r, column=16, value=url)
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=11, color="FF1E3D4C", underline="single")
        else:
            ws.cell(row=r, column=16, value="(not yet built)")
            ws.cell(row=r, column=16).font = Font(name="Calibri", size=11, italic=True, color="FF5C6970")

        # Spaced recall deep-link
        sr_url = BASE_URL + "bio304-spaced-recall-prototype.html#topic=" + tid
        cell = ws.cell(row=r, column=17, value=sr_url)
        cell.hyperlink = sr_url
        cell.font = Font(name="Calibri", size=11, color="FF1E3D4C", underline="single")

        # Lab workbook URL: workbook_day{day:02d}_{slugified-title}.html
        day = topic.get("day")
        if day:
            slug = slugify(topic["title"])
            wb_filename = f"workbook_day{day:02d}_{slug}.html"
            wb_url = BASE_URL + wb_filename
            cell = ws.cell(row=r, column=18, value=wb_url)
            cell.hyperlink = wb_url
            cell.font = Font(name="Calibri", size=11, color="FF1E3D4C", underline="single")

        filled += 1

    # Widen the new columns
    ws.column_dimensions["P"].width = 42
    ws.column_dimensions["Q"].width = 56
    ws.column_dimensions["R"].width = 56

    # Update the instructions row to reflect the new columns
    instr = ws.cell(row=3, column=1).value or ""
    new_instr = ("Paste video URLs in column L, note URLs in column M. Column N has the "
                 "OpenStax reference. Columns P, Q, R hold the lecture page, spaced-recall "
                 "deep-link, and lab workbook URLs respectively. Pre-work nights: Mon, Tue, "
                 "Thu, Fri. Wednesday is recall + lab. Topic ID in column C is what the "
                 "engine matches on.")
    if instr != new_instr:
        ws.cell(row=3, column=1, value=new_instr)

    wb.save(XLSX)
    print(f"Filled {filled} rows with URLs; skipped {skipped} unrecognized rows.")
    print(f"  Output: {XLSX}")


if __name__ == "__main__":
    main()
